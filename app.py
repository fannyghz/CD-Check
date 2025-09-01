# import streamlit as st
# import easyocr
# import cv2
# import re
# import os
# import pandas as pd
# from tempfile import TemporaryDirectory
# from openpyxl import load_workbook
# import io

# # --- knobs ---
# UPSCALE_X = 2.0
# UPSCALE_Y = 2.0
# # -------------

# def save_to_template(records, template_path="blank xlsx.xlsx"):
#     # Load your blank Excel template
#     wb = load_workbook(template_path)
#     ws = wb.active

#     # Fill values in columns C & D starting at row 3
#     row = 3
#     col_c, col_d = "C", "D"
#     for i in range(0, len(records), 2):
#         if i < len(records):
#             ws[f"{col_c}{row}"] = records[i]["measurement"]
#         if i + 1 < len(records):
#             ws[f"{col_d}{row}"] = records[i+1]["measurement"]
#         row += 1

#     # Save workbook to a BytesIO buffer
#     out = io.BytesIO()
#     wb.save(out)
#     out.seek(0)
#     return out

# def extract_measurement(results):
#     for i, (bbox, text, conf) in enumerate(results):
#         m = re.search(r'(\d+(?:\.\d+)?)\s*nm', text.lower())
#         if m:
#             try:
#                 return float(m.group(1))
#             except ValueError:
#                 return None
#             #return m.group(1)

#         m_num = re.search(r'\d+(?:\.\d+)?', text)
#         if m_num:
#             num = m_num.group(0)
#             if i + 1 < len(results) and "nm" in results[i + 1][1].lower():
#                 return num
#             if (i + 2) < len(results):
#                 next_text = results[i+1][1]
#                 nextnext_text = results[i+2][1].lower()
#                 if re.fullmatch(r"\d", next_text) and "nm" in nextnext_text:
#                     return f"{num}.{next_text}"
#     return None

# def process_image(reader, img_bytes):
#     file_bytes = bytearray(img_bytes.read())
#     img = cv2.imdecode(np.frombuffer(file_bytes, np.uint8), cv2.IMREAD_COLOR)

#     h, w = img.shape[:2]
#     x1, x2 = int(w * 0.3), int(w * 0.7)
#     y1, y2 = int(h * 0.3), int(h * 0.7)
#     roi = img[y1:y2, x1:x2]

#     roi = cv2.resize(roi, None, fx=UPSCALE_X, fy=UPSCALE_Y, interpolation=cv2.INTER_LANCZOS4)

#     results = reader.readtext(
#         roi,
#         allowlist="0123456789.nm",
#         detail=1,
#         mag_ratio=1.5,
#         add_margin=0.05,
#     )
#     return extract_measurement(results)


# def main():
#     st.title("CD Measurement Extractor")
#     st.write("Upload `.tif` images and extract measurements automatically.")

#     uploaded_files = st.file_uploader("Upload images", type=["tif", "tiff"], accept_multiple_files=True)

#     if uploaded_files:
#         reader = easyocr.Reader(["en"])
#         records = []
#         for file in uploaded_files:
#             measurement = process_image(reader, file)
#             records.append({"file": file.name, "measurement": measurement})

#         df = pd.DataFrame(records)
#         st.dataframe(df)

#         # Save to Excel
#         if st.button("Download Excel"):
#             filled_file = save_to_template(records, template_path="Blank_CD_Table.xlsx")
#             st.download_button(
#                 "Download Results",
#                 data=filled_file,
#                 file_name="CD Check.xlsx",
#                 mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#             )

# if __name__ == "__main__":
#     import numpy as np
#     main()



### VERSION 2 ###


# import streamlit as st
# import pandas as pd
# import easyocr
# import re
# import cv2
# import os
# import io
# from openpyxl import load_workbook

# # --- knobs ---
# UPSCALE_X = 2.0
# UPSCALE_Y = 2.0
# TEMPLATE_PATH = "Blank_CD_Table.xlsx"   # local template file
# # -------------

# def extract_measurement(results):
#     """
#     Extract 'xxx.x nm' from EasyOCR results, with fix for
#     cases like '236' + '7' + 'nm'.
#     """
#     for i, (bbox, text, conf) in enumerate(results):
#         # Case A: number and 'nm' in same chunk
#         m = re.search(r'(\d+(?:\.\d+)?)\s*nm', text.lower())
#         if m:
#             return m.group(1)

#         # Case B: number in this chunk
#         m_num = re.search(r'\d+(?:\.\d+)?', text)
#         if m_num:
#             num = m_num.group(0)

#             # B1: next chunk is 'nm'
#             if i + 1 < len(results) and 'nm' in results[i+1][1].lower():
#                 return num

#             # B2: decimal split
#             if (i + 2) < len(results):
#                 next_text = results[i+1][1]
#                 nextnext_text = results[i+2][1].lower()
#                 if re.fullmatch(r'\d', next_text) and 'nm' in nextnext_text:
#                     return f"{num}.{next_text}"
#     return None


# def process_image(reader, image_path):
#     img = cv2.imread(image_path)
#     h, w = img.shape[:2]

#     # Central 40% crop
#     x1, x2 = int(w * 0.3), int(w * 0.7)
#     y1, y2 = int(h * 0.3), int(h * 0.7)
#     roi = img[y1:y2, x1:x2]

#     # Upscale ROI
#     roi = cv2.resize(roi, None, fx=UPSCALE_X, fy=UPSCALE_Y, interpolation=cv2.INTER_LANCZOS4)

#     # OCR
#     results = reader.readtext(
#         roi,
#         allowlist='0123456789.nm',
#         detail=1,
#         mag_ratio=1.5,
#         add_margin=0.05
#     )

#     return extract_measurement(results)


# def main():
#     st.title("Auto CD Measurement Extractor")

#     uploaded_files = st.file_uploader(
#         "Upload your .tif images",
#         type=["tif", "tiff"],
#         accept_multiple_files=True
#     )

#     if uploaded_files:
#         if st.button("🔍 Extract Measurements"):
#             reader = easyocr.Reader(["en"])

#             # Progress bar
#             progress = st.progress(0)
#             records = []
#             n_files = len(uploaded_files)

#             for i, uploaded_file in enumerate(uploaded_files):
#                 # Save temp file
#                 temp_path = f"temp_{uploaded_file.name}"
#                 with open(temp_path, "wb") as f:
#                     f.write(uploaded_file.read())

#                 measurement = process_image(reader, temp_path)
#                 records.append(measurement)

#                 progress.progress((i + 1) / n_files)
#                 os.remove(temp_path)


#             # --- Fill into fixed template ---
#             in_mem_file = io.BytesIO()
#             wb = load_workbook(TEMPLATE_PATH)
#             ws = wb.active

#             row = 3
#             col = 3  # column C
#             for val in records:
#                 ws.cell(row=row, column=col, value=float(val) if val else None)
#                 if col == 3:
#                     col = 4
#                 else:
#                     col = 3
#                     row += 1

#             wb.save(in_mem_file)
#             in_mem_file.seek(0)

#             st.download_button(
#                 "📥 Download Measurements",
#                 in_mem_file,
#                 file_name="CD_Check.xlsx",
#                 mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#             )


# if __name__ == "__main__":
#     main()


### VERSION 3 ###

import streamlit as st
import easyocr
import re
import cv2
import os
import io
from openpyxl import load_workbook

# --- knobs ---
UPSCALE_X = 2.0
UPSCALE_Y = 2.0
TEMPLATE_PATH = "Blank_CD_Table.xlsx"   # local template file
# -------------

def extract_measurement(results):
    """
    Extract 'xxx.x nm' from EasyOCR results, with fix for
    cases like '236' + '7' + 'nm'.
    """
    for i, (bbox, text, conf) in enumerate(results):
        # Case A: number and 'nm' in same chunk
        m = re.search(r'(\d+(?:\.\d+)?)\s*nm', text.lower())
        if m:
            return m.group(1)

        # Case B: number in this chunk
        m_num = re.search(r'\d+(?:\.\d+)?', text)
        if m_num:
            num = m_num.group(0)

            # B1: next chunk is 'nm'
            if i + 1 < len(results) and 'nm' in results[i+1][1].lower():
                return num

            # B2: decimal split
            if (i + 2) < len(results):
                next_text = results[i+1][1]
                nextnext_text = results[i+2][1].lower()
                if re.fullmatch(r'\d', next_text) and 'nm' in nextnext_text:
                    return f"{num}.{next_text}"
    return None


def process_image(reader, image_path):
    img = cv2.imread(image_path)
    h, w = img.shape[:2]

    # Central 40% crop
    x1, x2 = int(w * 0.3), int(w * 0.7)
    y1, y2 = int(h * 0.3), int(h * 0.7)
    roi = img[y1:y2, x1:x2]

    # Upscale ROI
    roi = cv2.resize(roi, None, fx=UPSCALE_X, fy=UPSCALE_Y, interpolation=cv2.INTER_LANCZOS4)

    # OCR
    results = reader.readtext(
        roi,
        allowlist='0123456789.nm',
        detail=1,
        mag_ratio=1.5,
        add_margin=0.05
    )

    return extract_measurement(results)


def main():
    st.title("📏 AutoCD - Measure Extractor")
    st.markdown("### Upload `.tif` images")
    st.caption("Drag & drop multiple `.tif` images here. Only `.tif` files will be considered.")

    uploaded_files = st.file_uploader(
        "Select your images",
        type=["tif", "tiff"],
        accept_multiple_files=True
    )

    if uploaded_files:
        if st.button("🚀 Start Extraction"):
            reader = easyocr.Reader(["en"])

            # Progress bar
            progress = st.progress(0)
            records = []
            n_files = len(uploaded_files)

            for i, uploaded_file in enumerate(uploaded_files):
                temp_path = f"temp_{uploaded_file.name}"
                with open(temp_path, "wb") as f:
                    f.write(uploaded_file.read())

                measurement = process_image(reader, temp_path)
                records.append(measurement)

                progress.progress((i + 1) / n_files)
                os.remove(temp_path)

            # --- Fill into fixed template ---
            in_mem_file = io.BytesIO()
            wb = load_workbook(TEMPLATE_PATH)
            ws = wb.active

            row = 3
            col = 3  # column C
            for val in records:
                ws.cell(row=row, column=col, value=float(val) if val else None)
                if col == 3:
                    col = 4
                else:
                    col = 3
                    row += 1

            wb.save(in_mem_file)
            in_mem_file.seek(0)

            st.success("✅ Extraction complete!")
            st.download_button(
                "📥 Download Results",
                in_mem_file,
                file_name="CD_Check.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )


if __name__ == "__main__":
    main()
